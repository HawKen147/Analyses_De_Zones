from openpyxl import *
from datetime import datetime
import os
from openpyxl.styles import Font, Border, PatternFill, Alignment, Side

#pip install pillow -> inserer des images dans un fichier excel


wb = Workbook()

ws = wb.active

ws.title = f"Essaies zones {datetime.today().strftime('%d-%m-%Y')}"


#Nom du fichier pour qu'il soit sauvegardé
file_name = f"Essaies_zones_{datetime.today().strftime('%d-%m-%Y')}.xlsx"

i = 0
while True:
    i += 1
    if not os.path.isfile(f"./excl-test/{file_name}"):
            break 
    else:
        file_name = f"Essaies_zones_{datetime.today().strftime('%d-%m-%Y')}({i}).xlsx"

#Definit la largeur des colonnes
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 18
#Format la largeur des colonnes en fonctions de si elles sont pairs ou impairs
number_odd = 1
for cols in ws.iter_cols(min_col=3, max_col=20):
    #Obtient la lettre de la colonne à partir de la première cellule de la colonne
    col_letter = cols[0].column_letter
    if number_odd % 2 == 1:
        ws.column_dimensions[col_letter].width = 26
    else:
        ws.column_dimensions[col_letter].width = 15
    number_odd += 1

#Définit la hauteur des lignes
ws.row_dimensions[1].height = 64.5
ws.row_dimensions[2].height = 33
ws.row_dimensions[4].height = 18

#Fusion de toutes les cellules
ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=2) #Carré vide en haut a gauche du doc
ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4) # RTE -
ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=17) # Poste de xxxx \n tableau essaie zone \n xxxx_rapport_essaie_zone
ws.merge_cells(start_row=1, start_column=18, end_row=2, end_column=20) #Logo SPIE
ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=6) #Créer le xx/xx/xxxx \n auteur de création Pierry BENOIT
ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=14) #Modifier le : \n acteur de modification :
ws.merge_cells(start_row=2, start_column=15, end_row=2, end_column=17) #Version : 1
ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=8) #Début de zone
ws.merge_cells(start_row=3, start_column=9, end_row=3, end_column=14) #Milieux de zone
ws.merge_cells(start_row=3, start_column=15, end_row=3, end_column=20) #Fin de zone

#Rempli les cellules fusionnées
ws.cell(row=1, column=3).value = 'RTE -'
ws.cell(row=1, column=5).value = 'Poste de XXXX \nTableau des essaies de zone \nXXXX_Rapport_Essais_Zone'
ws.cell(row=1, column=18).value = 'Logo de SPIE'
ws.cell(row=2, column=3).value = f'Créer le : {datetime.today().strftime('%d/%m/%Y')} \nAuteur de la création : Pierry BENOIT'
ws.cell(row=2, column=7).value = "Modifié le :  \nActeur de la modification : "
ws.cell(row=2, column=15).value = "Version : 1"
ws.cell(row=3, column=3).value = "Début de zone"
ws.cell(row=3, column=9).value = "Milieu de zone"
ws.cell(row=3, column=15).value = "Fin de zone"

#Rempli les cellules non fusionnées
ws['A4'] = "N° de zone"
ws['B4'] = "TH associé"
ws['C4'] = "Début Marché"
ws['D4'] = "Exp."
ws['E4'] = "Début courir"
ws['F4'] = "Exp."
ws['G4'] = "Début rampé"
ws['H4'] = "Exp."
ws['I4'] = "Milieu marché"
ws['J4'] = "Exp."
ws['K4'] = "Milieu courir"
ws['L4'] = "Exp."
ws['M4'] = "Milieu rampé"
ws['N4'] = "Exp."
ws['O4'] = "Fin marché"
ws['P4'] = "Exp."
ws['Q4'] = "Fin courir"
ws['R4'] = "Exp."
ws['S4'] = "Fin rampé"
ws['T4'] = "Exp."

#Applique un retour a la ligne aux cellules
retour_ligne = Alignment(wrapText=True)
for cells in ws[1:2]:
        for cell in cells :
                cell.alignment = retour_ligne

#Mise en page du fichier excel
##Met le texte au centre de la cellule
texte_centrer = Alignment(horizontal='center', vertical='center')
ws.cell(row=1, column=3).alignment = texte_centrer
ws.cell(row=1, column=5).alignment = texte_centrer
for cells in ws[3:4]:
    for cell in cells :
        cell.alignment = texte_centrer

##Change l'arrière plan de la cellule (couleurs)
couleur_cell_C3 = "f7d059"
couleur_cell_I3 = "3bf77a"
couleur_cell_O3 = "fff700"
couleur_row_4 = "59a0f7"
ws['C3'].fill = PatternFill(start_color=couleur_cell_C3, end_color=couleur_cell_C3, fill_type="solid") #Change le fond de la cellule pour du vert
ws['I3'].fill = PatternFill(start_color=couleur_cell_I3, end_color=couleur_cell_I3, fill_type="solid") #Change le fond de la cellule pour du jaune
ws['O3'].fill = PatternFill(start_color=couleur_cell_O3, end_color=couleur_cell_O3, fill_type="solid") #Change le fond de la cellule pour du bleu
for cell in ws[4]:
     cell.fill = PatternFill(start_color=couleur_row_4, end_color=couleur_row_4, fill_type="solid") #Change le fond de la cellule en bleu

##Met le texte en gras 
for cells in ws[1:3]:
     for cell in cells:
          cell.font = Font(bold=True)

##Met les bords de la cellule en gras
medium = Side(border_style="medium", color="000000" )
border_style = Border (left=medium, right=medium, top=medium, bottom=medium)
for cells in ws[1:4]:
     for cell in cells:
          cell.border = border_style

##Change la taille de la police 
ws['C1'].font = Font(size=20, bold=True)
ws['E1'].font = Font(size=14, bold=True)
for cell in ws[4]:
     cell.font = Font(size=14, bold=True, color="ffffff")

##Définition des styles pour la mise en forme du fichier excel
pair_lines_background = PatternFill(start_color='adcdff', end_color='adcdff', fill_type="solid")
border_zone_TH = Border(left=Side(border_style="thick",color='FF000000'),
                        right=Side(border_style="thick",color='FF000000'),
                        vertical=Side(border_style="thick",color='FF000000'),
                        bottom=Side(border_style="dotted",color='FF000000'))
border_exp = Border(right=Side(border_style='double',color='FF000000'),
                            bottom=Side(border_style='dotted',color='FF000000'))
border_passage = Border(right=Side(border_style='dotted',color='FF000000'),
                            bottom=Side(border_style='dotted',color='FF000000'))
border_top =  Border(top=Side(border_style="medium", color='FF000000'))

##Application des styles sur les lignes des passages
nb_camera = 20
for rows in ws.iter_rows(min_row=5, max_row=nb_camera, min_col=1, max_col=20):
    for cell in rows:
        if cell.column_letter == 'A' or cell.column_letter == 'B' :
            cell.border = border_zone_TH
        elif cell.column % 2 == 0:
            cell.border = border_exp 
        else : 
            cell.border = border_passage
            
        if (cell.row % 2 == 0):
            cell.fill = pair_lines_background

##Applique le style a la dernier ligne +1
for cell in ws.iter_rows(min_row=nb_camera + 1, max_row=nb_camera + 1, min_col=1, max_col=20):
    for cell_in_row in cell:
        cell_in_row.border = border_top


#Sauvegarde le weebook (fichier excel)
wb.save(f"./excl-test/{file_name}")