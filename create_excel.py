from openpyxl import load_workbook
import fonctions as fonctions
import datetime
import locale
import os

def parcour_num_cam_excel (sheet):
    # Parcourir la colonne B (index 2) à partir de la ligne 2 (en supposant que les en-têtes sont à la ligne 4)
    for cell in sheet.iter_rows(min_row=5, min_col=2, max_col=2):
        # Accéder à la valeur de la cellule actuelle
        valeur_cellule = cell[0].value  # Fonctionne si la cellule n'est pas vide
        print(valeur_cellule)

#definit la bonne colonne de la feuille excel selon le type de passage
#return la lettre de la colonne de la feuille excel
def colonne_passage_excel(type_passage):
    match type_passage:
        case "DM" :
            return 'C'
        case "DC":
            return 'E'
        case "DR":
            return 'G'
        case "MM":
            return 'I'
        case "MC":
            return 'K'
        case "MR" :
            return 'M'
        case "FM" :
            return 'O'
        case "FC" :
            return 'Q'
        case "FR" :
            return 'S'
        case _:
            print(f"le type de passage rejeté {type_passage}")
            return False
        
#Récupere l'heure du passage du clip
#retourne la date et l'heure du passage
def recuperer_heure_passage(nom_clip):
    for i in range(len(nom_clip)):
        if nom_clip[i:i+4].isdigit():
            time = nom_clip[i:i+16]
            date, heure = time.split('_')
            return date, heure
        
#formate la date YYYY-MM-DD en jour xx mois année (vendredi 17 Mai 2024)
#retourne la date formaté
def format_date(date):
    date_obj = datetime.datetime.strptime(date, "%Y-%m-%d") 
    #locale.setlocale(locale.LC_ALL, 'fr_FR')                  #Transforme la date en francais (au lieu d'anglais) ############### Cette ligne fait bugger tout le script et ne dois pas etre utilisé
    formatted_date = date_obj.strftime("%A %d %B %Y")
    return formatted_date

#fonction qui verifie si le dossier pour stoker le fichier excel existe ou pas
#si il n'existe pas il crée le dossier
def check_path():
    if not "final" in os.listdir(f"model/excel"):
        os.makedirs(f"excel/final")

#Fonction principale de la création du fichier excel
def main_excel(path_to_folder):

    #fonction qui verifie si le dossier pour stoker le fichier excel existe ou pas
    #si il n'existe pas il crée le dossier
    check_path()

    # Chemin du fichier Excel final
    new_excel_file = "./model/excel/final/Essaies_Zones.xlsx"

    # Charge le classeur Excel existant
    workbook = load_workbook("./model/excel/model/model.xlsx")

    # Sélectionne la feuille active
    sheet = workbook.active

    # Tant que le fichier excel existe, on change le nom du fichier en ajoutant l'indice au nombre de fichier
    i = 0
    while True:
        i += 1
        if not os.path.isfile(new_excel_file):
            break 
        else:
            new_excel_file = f"./model/excel/final/Essaies_Zones({i}).xlsx"

    path_to_folder = "C:/Users/pierry.benoit/Documents/dossiers_camera"         #Il faut récuperer le chemin !

    # Obtenir le nombre de caméras
    nb_camera = fonctions.get_nb_camera(path_to_folder)

    # Insérer le numero de camera dans la bonne cellule
    for i in range(1, nb_camera + 1):
        sheet[f'B{i + 4}'] = i

    #recupere tous les clips videos dans un tableau [[clip_ramper_1, clip_ramper_2, clip_ramper_3], [clip_marcher_1, clip_marcher_2, clip_marcher_3]...]
    clips = fonctions.get_videos_clips(path_to_folder)

    #On parcourt tout les clips
    for clip in clips :
        index_numero_cam = clip.find("TH")                              #Index ou se situe TH (récuperer l'index de T)
        numero_cam = clip[index_numero_cam+2:index_numero_cam+4]        #On recupere l'indice ou se situe TH et on récupere les 2 chiffres apres TH
        ligne_excel = int(numero_cam) + 4                                   #On commence a la 4è ligne du fichier excel (apres les en-tetes)
        type_de_passage = clip[-6:-4]                                   #On récupere le type de passage (qui est la fin du nom de la video ...DM.asf)
        colonne_excel = colonne_passage_excel(type_de_passage)              #on récuperer la bonne colonne en fonction du passage (DM = colonne C)
        date, heure = recuperer_heure_passage(clip)
        heure = heure.replace('h', ':')                                     #On remplace HHhMM en HH:MM (14h00 -> 14:00)
        date_time = format_date(date)                                       #Formate la date en Francais
        new_date_time = fonctions.date_eng_to_fr(date_time)
        sheet[f"{colonne_excel}{ligne_excel}"] = f"{new_date_time} {heure}"          #Remplis la feuille de calcule

    #On enregistre le classeur Excel
    workbook.save(new_excel_file)
